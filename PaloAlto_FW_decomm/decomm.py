import openpyxl
import datetime
import re

# Input / output filenames
INPUT_EXCEL = 'rules_to_decommission.xlsx'
OUTPUT_TXT = 'firewall_commands.txt'

# Load workbook and pick active sheet
wb = openpyxl.load_workbook(INPUT_EXCEL)
sheet = wb.active

# Map header names (lowercase) to column indices
header = {}
for col in range(1, sheet.max_column + 1):
    cell_val = sheet.cell(row=1, column=col).value
    if cell_val is not None:
        header[cell_val.strip().lower()] = col

# Read rule entries
rules = []
for row in range(2, sheet.max_row + 1):
    rules.append({
        'vsys': sheet.cell(row, header['vsys']).value,
        'rule_name': sheet.cell(row, header['rule_name']).value,
        'change_name': sheet.cell(row, header['change_name']).value,
        'implementer': sheet.cell(row, header['implementer']).value,
    })

# Current date in dd.mm.YYYY
current_date = datetime.date.today().strftime('%d.%m.%Y')

# Containers for commands
tag_cmds = []
desc_cmds = []
disable_cmds = []
move_cmds = []

def build_cmd(parts):
    """
    Join the parts list with a single space, drop any empty parts,
    and collapse runs of multiple spaces into a single space.
    """
    # Drop empty or None parts
    parts = [p for p in parts if p]
    # Join with single spaces
    cmd = " ".join(parts)
    # Collapse multiple spaces just in case
    cmd = re.sub(r" {2,}", " ", cmd)
    return cmd

for rule in rules:
    vsys = rule.get('vsys')
    rule_name = rule.get('rule_name')
    change_name = rule.get('change_name')
    implementer = rule.get('implementer')

    # Convert to string and strip whitespace
    vsys = (str(vsys) if vsys is not None else "").strip()
    rule_name = (str(rule_name) if rule_name is not None else "").strip()
    change_name = (str(change_name) if change_name is not None else "").strip()
    implementer = (str(implementer) if implementer is not None else "").strip()

    # Debug print to inspect contents (optional)
    print("DEBUG vsys:", repr(vsys),
          "rule_name:", repr(rule_name),
          "change_name:", repr(change_name),
          "implementer:", repr(implementer))

    # Skip if rule_name is empty
    if not rule_name:
        continue

    # Defaults
    if not change_name:
        change_name = "NoChange"
    if not implementer:
        implementer = "Unknown"

    description = f"{change_name} / {current_date} {implementer}"

    # Build tag command
    tag_cmds.append(build_cmd([
        "set", "device-group", vsys,
        "post-rulebase", "security", "rules", rule_name,
        "tag", f"{change_name}_decomm"
    ]))

    # Build description command
    desc_cmds.append(build_cmd([
        "set", "device-group", vsys,
        "post-rulebase", "security", "rules", rule_name,
        "description", f"\"{description}\""
    ]))

    # Build disable command
    disable_cmds.append(build_cmd([
        "set", "device-group", vsys,
        "post-rulebase", "security", "rules", rule_name,
        "disable", "yes"
    ]))

    # Build move command
    move_cmds.append(build_cmd([
        "move", "device-group", vsys,
        "rulebase", "security", "rules", rule_name,
        "bottom"
    ]))

# Combine all the commands with blank lines between the groups
all_cmds = (
    tag_cmds +
    [""] +
    desc_cmds +
    [""] +
    disable_cmds +
    [""] +
    move_cmds
)

# Write to output file
with open(OUTPUT_TXT, 'w') as outf:
    for ln in all_cmds:
        outf.write(ln + "\n")

print(f"Commands written to {OUTPUT_TXT}")