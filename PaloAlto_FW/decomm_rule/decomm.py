import openpyxl
import datetime

# Input Excel filename
INPUT_EXCEL = 'rules_to_decommission.xlsx'
# Output TXT filename with CLI commands
OUTPUT_TXT = 'firewall_commands.txt'

# Load workbook and select active sheet
wb = openpyxl.load_workbook(INPUT_EXCEL)
sheet = wb.active  # Change if needed, e.g., wb['Sheet1']

# Read headers to map columns
header = {}
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=1, column=col).value
    if cell_value:
        header[cell_value.strip().lower()] = col

# Read rules data
rules = []
for row in range(2, sheet.max_row + 1):
    rules.append({
        'vsys': sheet.cell(row, header['vsys']).value,
        'rule_name': sheet.cell(row, header['rule_name']).value,
        'change_name': sheet.cell(row, header['change_name']).value,
        'implementer': sheet.cell(row, header['implementer']).value,
    })

# Get current date in dd.mm.YYYY format
current_date = datetime.date.today().strftime('%d.%m.%Y')

# Prepare command lists by type
tag_commands = []
description_commands = []
disable_commands = []
move_bottom_commands = []

# Process each rule
for rule in rules:
    vsys = str(rule['vsys']).strip() if rule['vsys'] else ''
    rule_name = str(rule['rule_name']).strip() if rule['rule_name'] else ''
    change_name = str(rule['change_name']).strip() if rule['change_name'] else ''
    implementer = str(rule['implementer']).strip() if rule['implementer'] else ''

    # Skip rules with empty rule_name
    if not rule_name:
        continue

    # Default values if change_name or implementer is missing
    if not change_name:
        change_name = "NoChange"
    if not implementer:
        implementer = "Unknown"

    # Generate description
    description = f"{change_name} / {current_date} {implementer}"

    # 1. Add tag command (with vsys)
    tag_commands.append(f"set device-group {vsys} post-rulebase security rules {rule_name} tag {change_name}_decomm")

    # 2. Update description (with vsys)
    description_commands.append(f"set device-group {vsys} post-rulebase security rules {rule_name} description \"{description}\"")

    # 3. Disable the rule (with vsys)
    disable_commands.append(f"set device-group {vsys} post-rulebase security rules {rule_name} disable yes")

    # 4. Move rule to bottom (with vsys)
    move_bottom_commands.append(f"move device-group {vsys} rulebase security rules {rule_name} bottom")

# Combine all commands in the desired order (tag, description, disable, move bottom)
all_commands = tag_commands + [''] + description_commands + [''] + disable_commands + [''] + move_bottom_commands

# Write all commands to output file
with open(OUTPUT_TXT, 'w') as outfile:
    for line in all_commands:
        outfile.write(line + '\n')

print(f"Commands written to {OUTPUT_TXT}")
